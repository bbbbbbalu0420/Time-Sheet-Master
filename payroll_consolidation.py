#!/usr/bin/env python3
"""
ММ РАСЧЕТ ГРАФИКА — Скрипт консолидации зарплатных данных
Собирает данные из отчётов по объектам и обновляет Мастер-файл.
"""

import os
import re
import sys
import glob
import shutil
import datetime

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# КОНСТАНТЫ
# ─────────────────────────────────────────────────────────────────────────────

SALARY_BASE = 5000.0          # Оклад (руб.)
SALARY_LIMIT = 24_500.0       # Максимальная ЗП в месяц
TARGET_HOURS_MIN = 350        # Целевой минимум часов
TARGET_HOURS_MAX = 375        # Целевой максимум часов

# Производственный календарь 2026 (40-ч. неделя, нормы часов)
WORK_HOURS_NORM_2026 = {
    1: 136,  # Январь
    2: 151,  # Февраль
    3: 167,  # Март
    4: 175,  # Апрель
    5: 143,  # Май
    6: 167,  # Июнь
    7: 184,  # Июль
    8: 168,  # Август
    9: 176,  # Сентябрь
    10: 176, # Октябрь
    11: 159, # Ноябрь
    12: 175, # Декабрь
}

MONTH_NAMES_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 1 — Запрос месяца у пользователя
# ─────────────────────────────────────────────────────────────────────────────

def ask_month() -> int:
    """Запрашивает номер месяца у пользователя."""
    print("\n" + "="*60)
    print("  ММ РАСЧЕТ ГРАФИКА — Консолидация зарплатных данных")
    print("="*60)
    print("\nМесяцы 2026 года:")
    for num, name in MONTH_NAMES_RU.items():
        norm = WORK_HOURS_NORM_2026[num]
        print(f"  {num:2d}. {name:<12} (норма {norm} ч.)")

    while True:
        try:
            month = int(input("\nВведите номер месяца (1-12): ").strip())
            if 1 <= month <= 12:
                return month
            print("  Ошибка: введите число от 1 до 12")
        except ValueError:
            print("  Ошибка: введите целое число")


# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 2 — Чтение Мастер-файла
# ─────────────────────────────────────────────────────────────────────────────

def parse_master_file(master_path: str, month: int):
    """
    Читает Мастер-файл и возвращает:
      - employees: dict {fio_clean -> {'row': int, 'status': str, 'dismiss_date': date|None,
                                       'day_cols': {day: col_index}}}
      - wb: workbook объект
      - ws: лист
      - header_row: номер строки с датами
      - day_col_map: {day_num: col_index} — первый столбец каждого дня
    """
    print(f"\n[1/5] Читаю Мастер-файл: {master_path}")
    wb = load_workbook(master_path, data_only=True)
    ws = wb.active

    # Ищем строку заголовка с числами дней (строка 3 в примере)
    header_row = None
    day_col_map = {}  # {day: first_col_index_1based}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        # Строка с днями месяца — содержит много целых чисел 1-31
        day_vals = [v for v in row if isinstance(v, (int, float)) and 1 <= v <= 31]
        if len(day_vals) >= 20:
            header_row = row_idx
            # Каждый день занимает 2 столбца (часы + ночные).
            # Записываем в первый столбец каждого дня.
            # Собираем все столбцы для каждого дня, берём первый (меньший).
            day_all_cols: dict = {}
            for col_idx, val in enumerate(row, 1):
                if isinstance(val, (int, float)) and 1 <= val <= 31:
                    day = int(val)
                    if day not in day_all_cols:
                        day_all_cols[day] = []
                    day_all_cols[day].append(col_idx)
            # Каждый день занимает 2 столбца: [дневные часы | ночные часы].
            # Часы записываются во второй столбец пары (нечётный/чётный зависит от начала).
            # На основе реального файла: значения стоят во ВТОРОМ столбце каждого дня.
            for day, cols in day_all_cols.items():
                if len(cols) >= 2:
                    day_col_map[day] = cols[1]  # Второй столбец дня (основные часы)
                else:
                    day_col_map[day] = cols[0]
            break

    if header_row is None:
        raise ValueError("Не найдена строка заголовка с датами в Мастер-файле!")

    print(f"  Строка заголовка найдена: строка {header_row}")
    print(f"  Найдено дней: {len(day_col_map)} ({min(day_col_map)} - {max(day_col_map)})")

    # Читаем сотрудников — ищем строки с ФИО (столбец B, начиная после заголовка)
    employees = {}
    fio_col = 2  # Столбец B (1-based)

    for row_idx in range(header_row + 2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=1).value
        cell_fio = ws.cell(row=row_idx, column=fio_col).value

        if cell_fio is None or not isinstance(cell_fio, str):
            continue
        cell_fio = cell_fio.strip()
        if not cell_fio:
            continue

        # Парсим статус из имени
        status, dismiss_date, sick_end = parse_employee_status(cell_fio)

        # Чистое ФИО (без статусных пометок после дефиса)
        fio_clean = extract_clean_fio(cell_fio)

        employees[fio_clean] = {
            'row': row_idx,
            'raw_fio': cell_fio,
            'status': status,
            'dismiss_date': dismiss_date,
            'sick_end': sick_end,
        }

    print(f"  Найдено сотрудников: {len(employees)}")
    for fio, info in employees.items():
        status_str = info['status']
        if info['dismiss_date']:
            status_str += f" (до {info['dismiss_date']})"
        print(f"    - {fio:<35} [{status_str}]")

    return employees, wb, ws, header_row, day_col_map


def extract_clean_fio(raw_fio: str) -> str:
    """Извлекает чистое ФИО без статусных пометок."""
    # Убираем всё после первого дефиса (статус объекта)
    parts = raw_fio.split('-')
    clean = parts[0].strip()
    return clean.upper()


def parse_employee_status(raw_fio: str):
    """
    Разбирает статус сотрудника из строки ФИО.
    Возвращает: (status, dismiss_date, sick_end_date)
    """
    raw_upper = raw_fio.upper()
    dismiss_date = None
    sick_end = None

    # Уволен — ищем дату увольнения
    m = re.search(r'УВОЛ[^\d]*(\d{1,2})[./](\d{2})[./](\d{2,4})', raw_upper)
    if m:
        day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        try:
            dismiss_date = datetime.date(yr, mon, day)
        except ValueError:
            pass
        return 'УВОЛЕН', dismiss_date, sick_end

    # Больничный — ищем дату конца
    m = re.search(r'БОЛН[^\d]*(?:ДО\s*)?(\d{1,2})[./](\d{2})[./](\d{2,4})', raw_upper)
    if m:
        day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        try:
            sick_end = datetime.date(yr, mon, day)
        except ValueError:
            pass
        return 'БОЛЬНИЧНЫЙ', dismiss_date, sick_end

    if 'БОЛН' in raw_upper:
        return 'БОЛЬНИЧНЫЙ', dismiss_date, sick_end

    # Отпуск
    if 'ОТПУСК' in raw_upper or 'ОТП' in raw_upper:
        return 'ОТПУСК', dismiss_date, sick_end

    return 'РАБОТАЕТ', dismiss_date, sick_end


# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 3 — Чтение отчётов по объектам
# ─────────────────────────────────────────────────────────────────────────────

def load_all_reports(objects_folder: str, month: int, year: int = 2026) -> dict:
    """
    Читает все файлы RESULT_PAYROLL_*.xlsx и *.csv из папки objects/.
    Возвращает dict: {fio_upper -> {day: hours}}
    """
    print(f"\n[2/5] Читаю отчёты из папки: {objects_folder}")

    # Поиск файлов
    xlsx_files = glob.glob(os.path.join(objects_folder, "RESULT_PAYROLL_*.xlsx"))
    xlsx_files += glob.glob(os.path.join(objects_folder, "RESULT_PAYROLL_*.xls"))
    csv_files = glob.glob(os.path.join(objects_folder, "RESULT_PAYROLL_*.csv"))
    all_files = xlsx_files + csv_files

    if not all_files:
        # Пробуем все xlsx в папке
        all_files = (
            glob.glob(os.path.join(objects_folder, "*.xlsx")) +
            glob.glob(os.path.join(objects_folder, "*.csv"))
        )

    if not all_files:
        print(f"  ПРЕДУПРЕЖДЕНИЕ: Файлы не найдены в папке {objects_folder}")
        return {}

    print(f"  Найдено файлов: {len(all_files)}")

    # Итоговый словарь: {fio -> {day -> hours}}
    all_hours: dict = {}

    for filepath in sorted(all_files):
        fname = os.path.basename(filepath)
        try:
            df = read_report_file(filepath)
            if df is None or df.empty:
                print(f"  Пропуск (пустой): {fname}")
                continue

            records = extract_hours_from_df(df, month, year)
            count = 0
            for fio, day_hours in records.items():
                if fio not in all_hours:
                    all_hours[fio] = {}
                for day, hrs in day_hours.items():
                    all_hours[fio][day] = all_hours[fio].get(day, 0) + hrs
                    count += 1
            print(f"  ✓ {fname:<50} — {len(records)} сотрудников, {count} записей")
        except Exception as e:
            print(f"  ✗ {fname}: ОШИБКА — {e}")

    return all_hours


def read_report_file(filepath: str) -> pd.DataFrame | None:
    """Читает файл отчёта (xlsx или csv)."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        for enc in ['utf-8', 'cp1251', 'utf-8-sig']:
            try:
                df = pd.read_csv(filepath, encoding=enc, sep=None, engine='python')
                return df
            except Exception:
                continue
        return None

    elif ext in ('.xlsx', '.xls'):
        try:
            xl = pd.ExcelFile(filepath)
            # Ищем лист "Ведомость" или берём первый
            sheet_name = xl.sheet_names[0]
            for s in xl.sheet_names:
                if 'ведомост' in s.lower() or 'payroll' in s.lower():
                    sheet_name = s
                    break
            df = xl.parse(sheet_name)
            return df
        except Exception as e:
            raise e
    return None


def extract_hours_from_df(df: pd.DataFrame, month: int, year: int) -> dict:
    """
    Извлекает часы из DataFrame отчёта.
    Ожидаемые колонки: 'Дата открытия', 'Кассир', 'НАЧИСЛЕНО'
    Возвращает {fio_upper -> {day -> hours}}
    """
    result = {}

    # Нормализуем названия столбцов
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.upper(): c for c in df.columns}

    # Ищем нужные столбцы
    date_col = None
    cashier_col = None
    hours_col = None

    for key, orig in col_map.items():
        if 'ДАТА' in key or 'DATE' in key:
            date_col = orig
        if 'КАССИР' in key or 'CASHIER' in key or 'ФИО' in key:
            cashier_col = orig
        if 'НАЧИСЛЕНО' in key or 'HOURS' in key or 'CHARGE' in key:
            hours_col = orig

    if not date_col or not cashier_col or not hours_col:
        missing = []
        if not date_col: missing.append("Дата открытия")
        if not cashier_col: missing.append("Кассир")
        if not hours_col: missing.append("НАЧИСЛЕНО")
        raise ValueError(f"Не найдены столбцы: {missing}")

    for _, row in df.iterrows():
        # Парсим дату
        raw_date = row[date_col]
        if pd.isna(raw_date):
            continue

        try:
            if isinstance(raw_date, (datetime.datetime, datetime.date)):
                dt = raw_date if isinstance(raw_date, datetime.datetime) else datetime.datetime.combine(raw_date, datetime.time())
            else:
                dt = pd.to_datetime(raw_date, dayfirst=True)
        except Exception:
            continue

        # Фильтр по нужному месяцу и году
        if dt.month != month or dt.year != year:
            continue

        day = dt.day

        # ФИО кассира
        fio = str(row[cashier_col]).strip().upper()
        if not fio or fio == 'NAN' or fio == 'СИС. АДМИНИСТРАТОР':
            continue

        # Часы
        hours_raw = row[hours_col]
        if pd.isna(hours_raw):
            continue
        try:
            hours = float(hours_raw)
        except (ValueError, TypeError):
            continue

        if hours <= 0:
            continue

        if fio not in result:
            result[fio] = {}
        result[fio][day] = result[fio].get(day, 0) + hours

    return result


# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 4 — Сопоставление ФИО из отчётов с Мастер-файлом
# ─────────────────────────────────────────────────────────────────────────────

def match_employees(report_data: dict, master_employees: dict) -> dict:
    """
    Сопоставляет ФИО из отчётов с ФИО в мастере.
    Возвращает {master_fio -> {day -> hours}}
    """
    print(f"\n[3/5] Сопоставляю сотрудников...")

    matched = {}
    unmatched_report = []

    # Готовим нормализованные ключи мастера
    master_keys = {}
    for fio in master_employees:
        normalized = normalize_fio(fio)
        master_keys[normalized] = fio

    for report_fio, day_hours in report_data.items():
        norm_report = normalize_fio(report_fio)

        # Точное совпадение
        if norm_report in master_keys:
            master_fio = master_keys[norm_report]
            matched[master_fio] = day_hours
            print(f"  ✓ Точное: {report_fio}")
            continue

        # Нечёткое совпадение — ищем по части ФИО
        best_match = fuzzy_match_fio(norm_report, master_keys)
        if best_match:
            master_fio = master_keys[best_match]
            matched[master_fio] = day_hours
            print(f"  ~ Нечёткое: {report_fio} → {master_fio}")
        else:
            unmatched_report.append(report_fio)

    if unmatched_report:
        print(f"\n  ВНИМАНИЕ! Не сопоставлены сотрудники из отчётов:")
        for fio in unmatched_report:
            print(f"    ? {fio}")

    return matched


def normalize_fio(fio: str) -> str:
    """Нормализует ФИО: верхний регистр, убирает лишнее."""
    fio = fio.upper().strip()
    # Убираем отчество если есть инициалы
    fio = re.sub(r'\s+', ' ', fio)
    return fio


def fuzzy_match_fio(norm_report: str, master_keys: dict) -> str | None:
    """Нечёткое совпадение по фамилии и первым буквам имени."""
    # Берём первое слово (фамилия)
    parts_report = norm_report.split()
    if not parts_report:
        return None

    surname_report = parts_report[0]
    initials_report = parts_report[1][:1] if len(parts_report) > 1 else ""

    best = None
    best_score = 0

    for norm_master in master_keys:
        parts_master = norm_master.split()
        if not parts_master:
            continue
        surname_master = parts_master[0]

        # Совпадение фамилии
        if surname_report == surname_master:
            score = 10
        elif surname_report in surname_master or surname_master in surname_report:
            score = 5
        else:
            # Проверяем первые 4 буквы
            min_len = min(len(surname_report), len(surname_master), 4)
            if min_len >= 3 and surname_report[:min_len] == surname_master[:min_len]:
                score = 3
            else:
                continue

        # Бонус за совпадение инициала
        if initials_report and len(parts_master) > 1:
            if parts_master[1][:1] == initials_report:
                score += 3

        if score > best_score:
            best_score = score
            best = norm_master

    return best if best_score >= 5 else None


# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 5 — Применение бизнес-правил и расчёт ЗП
# ─────────────────────────────────────────────────────────────────────────────

def apply_business_rules(
    matched_hours: dict,
    master_employees: dict,
    month: int,
    year: int,
) -> dict:
    """
    Применяет все бизнес-правила и рассчитывает итоговые часы.
    Возвращает {master_fio -> {day -> final_hours, 'salary': float, 'total_hours': float}}
    """
    print(f"\n[4/5] Применяю бизнес-правила...")

    norm_hours = WORK_HOURS_NORM_2026.get(month, 167)
    hour_cost = SALARY_BASE / norm_hours
    print(f"  Норма часов: {norm_hours} ч.")
    print(f"  Стоимость 1 часа: {hour_cost:.4f} руб.")

    results = {}

    # Определяем последний день месяца
    if month == 12:
        last_day = (datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)).day
    else:
        last_day = (datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)).day

    for master_fio, emp_info in master_employees.items():
        status = emp_info['status']
        dismiss_date = emp_info['dismiss_date']
        sick_end = emp_info['sick_end']

        # Получаем часы из отчётов
        raw_hours = matched_hours.get(master_fio, {})

        # ПРАВИЛО 1: Уволенные
        if status == 'УВОЛЕН' and dismiss_date:
            if dismiss_date.month == month and dismiss_date.year == year:
                if dismiss_date.day < 17:
                    # Часы = 0
                    print(f"  УВОЛЕН до 17-го: {master_fio} → 0 часов")
                    results[master_fio] = {
                        'day_hours': {},
                        'total_hours': 0,
                        'salary': 0,
                        'status': status,
                    }
                    continue
                else:
                    # Часы только до даты увольнения
                    cutoff = dismiss_date.day
                    raw_hours = {d: h for d, h in raw_hours.items() if d <= cutoff}
                    print(f"  УВОЛЕН с 17+: {master_fio} → учитываем по {cutoff} число")

        # ПРАВИЛО 2: Отпуск/больничный
        if status in ('ОТПУСК',):
            # Полностью пропускаем часы в дни отпуска (все дни месяца)
            print(f"  ОТПУСК: {master_fio} → часы не начисляем")
            results[master_fio] = {
                'day_hours': {},
                'total_hours': 0,
                'salary': 0,
                'status': status,
            }
            continue

        if status == 'БОЛЬНИЧНЫЙ':
            if sick_end:
                # Больничный только до sick_end
                if sick_end.month == month and sick_end.year == year:
                    cutoff = sick_end.day
                    raw_hours = {d: h for d, h in raw_hours.items() if d > cutoff}
                    print(f"  БОЛЬНИЧНЫЙ до {cutoff}: {master_fio} → часы только с {cutoff+1}")
                else:
                    # Больничный на весь месяц
                    raw_hours = {}
                    print(f"  БОЛЬНИЧНЫЙ весь месяц: {master_fio} → 0 часов")
            else:
                raw_hours = {}
                print(f"  БОЛЬНИЧНЫЙ: {master_fio} → 0 часов")

            if not raw_hours:
                results[master_fio] = {
                    'day_hours': {},
                    'total_hours': 0,
                    'salary': 0,
                    'status': status,
                }
                continue

        # Итоговые часы по дням
        day_hours = dict(raw_hours)
        total_hours = sum(day_hours.values())

        # РАСЧЁТ ЗП
        if total_hours <= norm_hours:
            # Только базовая оплата
            salary = total_hours * hour_cost
        else:
            # Переработка (сверх нормы) — двойная оплата
            overtime = total_hours - norm_hours
            salary = norm_hours * hour_cost + overtime * hour_cost * 2

        # ПРАВИЛО 3: Лимит ЗП 24 500 руб.
        if salary > SALARY_LIMIT:
            print(f"  ЛИМИТ: {master_fio} — ЗП {salary:.0f} руб. > {SALARY_LIMIT} руб. → обрезаем часы")
            # Находим максимально допустимые часы
            max_hours_base = SALARY_LIMIT / hour_cost  # часы при базовой оплате

            if max_hours_base <= norm_hours:
                # Лимит достигается в базовой части
                scale = SALARY_LIMIT / salary
                day_hours = {d: round(h * scale, 1) for d, h in day_hours.items()}
                total_hours = sum(day_hours.values())
                salary = SALARY_LIMIT
            else:
                # Лимит достигается в переработке
                # 5000 + (x - norm_hours) * 2 * hour_cost = 24500
                # x - norm_hours = (24500 - 5000) / (2 * hour_cost)
                max_overtime = (SALARY_LIMIT - SALARY_BASE) / (2 * hour_cost)
                max_total = norm_hours + max_overtime
                scale = max_total / total_hours
                day_hours = {d: round(h * scale, 1) for d, h in day_hours.items()}
                total_hours = sum(day_hours.values())
                salary = SALARY_LIMIT

        results[master_fio] = {
            'day_hours': day_hours,
            'total_hours': round(total_hours, 1),
            'salary': round(salary, 2),
            'status': status,
        }

    # Вывод сводки
    print("\n  Сводка расчётов:")
    print(f"  {'Сотрудник':<35} {'Статус':<12} {'Часы':>7} {'ЗП':>10}")
    print("  " + "-"*70)
    for fio, info in sorted(results.items()):
        print(f"  {fio:<35} {info['status']:<12} {info['total_hours']:>7.1f} {info['salary']:>10.2f}")

    return results


# ─────────────────────────────────────────────────────────────────────────────
# ШАГ 6 — Запись в Мастер-файл
# ─────────────────────────────────────────────────────────────────────────────

def write_to_master(
    wb: openpyxl.Workbook,
    ws,
    master_employees: dict,
    results: dict,
    day_col_map: dict,
    output_path: str,
):
    """
    Записывает итоговые часы в Мастер-файл и сохраняет как FINAL_MASTER_PAYROLL.xlsx
    """
    print(f"\n[5/5] Записываю результаты в Мастер-файл...")

    # Цвета для подсветки
    fill_red = PatternFill("solid", fgColor="FF0000")    # Уволен
    fill_yellow = PatternFill("solid", fgColor="FFFF00")  # Отпуск/больничный
    fill_green = PatternFill("solid", fgColor="90EE90")   # Норма
    fill_orange = PatternFill("solid", fgColor="FFA500")  # Переработка

    written_count = 0

    for master_fio, emp_info in master_employees.items():
        row_idx = emp_info['row']
        status = emp_info['status']

        if master_fio not in results:
            continue

        calc = results[master_fio]
        day_hours = calc['day_hours']

        # Определяем цвет строки по статусу
        if status == 'УВОЛЕН':
            row_fill = fill_red
        elif status in ('ОТПУСК', 'БОЛЬНИЧНЫЙ'):
            row_fill = fill_yellow
        else:
            row_fill = None

        # Записываем часы по дням
        for day, col in day_col_map.items():
            hours = day_hours.get(day, None)
            cell = ws.cell(row=row_idx, column=col)

            if hours is not None and hours > 0:
                cell.value = round(hours, 1)
                if row_fill:
                    cell.fill = row_fill
                written_count += 1
            elif hours == 0 or hours is None:
                if row_fill:
                    cell.fill = row_fill

        print(f"  ✓ {master_fio:<35} → {calc['total_hours']:>6.1f} ч. / {calc['salary']:>10.2f} руб.")

    print(f"\n  Записано ячеек: {written_count}")

    # Сохраняем файл
    wb.save(output_path)
    print(f"\n  Файл сохранён: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# ДОПОЛНИТЕЛЬНО — Создание сводного отчёта
# ─────────────────────────────────────────────────────────────────────────────

def create_summary_report(results: dict, month: int, year: int, output_dir: str):
    """Создаёт читаемый сводный отчёт в Excel."""
    summary_path = os.path.join(output_dir, f"СВОДКА_{MONTH_NAMES_RU[month]}_{year}.xlsx")

    rows = []
    for fio, calc in sorted(results.items()):
        rows.append({
            'ФИО': fio,
            'Статус': calc['status'],
            'Итого часов': calc['total_hours'],
            'Итого ЗП (руб.)': calc['salary'],
        })

    if rows:
        df = pd.DataFrame(rows)
        df.to_excel(summary_path, index=False)
        print(f"  Сводный отчёт: {summary_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Пути (можно изменить под свою структуру)
    OBJECTS_FOLDER = "objects"
    MASTER_FILE = "Master_Table.xlsx"
    OUTPUT_FILE = "FINAL_MASTER_PAYROLL.xlsx"
    YEAR = 2026

    # Шаг 1: Запрос месяца
    month = ask_month()
    month_name = MONTH_NAMES_RU[month]
    norm_hours = WORK_HOURS_NORM_2026[month]
    hour_cost = SALARY_BASE / norm_hours

    print(f"\n  Выбран: {month_name} {YEAR}")
    print(f"  Норма часов: {norm_hours} ч.")
    print(f"  Стоимость часа: {hour_cost:.2f} руб.")
    print(f"  Оклад: {SALARY_BASE:.0f} руб. / Лимит ЗП: {SALARY_LIMIT:.0f} руб.")

    # Проверяем пути
    if not os.path.exists(MASTER_FILE):
        # Ищем файл в текущей папке
        found = glob.glob("*.xlsx")
        if found:
            MASTER_FILE = found[0]
            print(f"\n  Мастер-файл не найден по умолчанию, использую: {MASTER_FILE}")
        else:
            print(f"\nОШИБКА: Мастер-файл '{MASTER_FILE}' не найден!")
            print("Убедитесь, что файл находится в текущей папке.")
            sys.exit(1)

    if not os.path.exists(OBJECTS_FOLDER):
        print(f"\nПАПКА '{OBJECTS_FOLDER}' НЕ НАЙДЕНА — создаю пустую папку.")
        print("Поместите файлы RESULT_PAYROLL_*.xlsx в папку 'objects/' и запустите снова.")
        os.makedirs(OBJECTS_FOLDER, exist_ok=True)

    # Копируем мастер-файл для редактирования
    shutil.copy2(MASTER_FILE, OUTPUT_FILE)

    # Шаг 2: Читаем Мастер-файл
    master_employees, wb, ws, header_row, day_col_map = parse_master_file(OUTPUT_FILE, month)

    # Шаг 3: Читаем отчёты объектов
    report_hours = load_all_reports(OBJECTS_FOLDER, month, YEAR)

    if not report_hours:
        print("\nВНИМАНИЕ: Данные из отчётов не загружены.")
        print(f"Проверьте, что в папке '{OBJECTS_FOLDER}' есть файлы RESULT_PAYROLL_*.xlsx")
        print("Скрипт создаст файл без изменений часов.")

    # Шаг 4: Сопоставляем сотрудников
    matched_hours = match_employees(report_hours, master_employees)

    # Шаг 5: Применяем бизнес-правила и считаем ЗП
    results = apply_business_rules(matched_hours, master_employees, month, YEAR)

    # Шаг 6: Записываем в файл
    write_to_master(wb, ws, master_employees, results, day_col_map, OUTPUT_FILE)

    # Создаём сводку
    create_summary_report(results, month, YEAR, ".")

    print("\n" + "="*60)
    print(f"  ГОТОВО! Файл сохранён: {OUTPUT_FILE}")
    print("="*60)


if __name__ == "__main__":
    main()
